from AudioTransformer import AudioClassifier
import openpyxl
from openpyxl import Workbook
import numpy as np
import torch.nn as nn
from lightning.pytorch import seed_everything
import argparse
from dataset import AudioDataset
from pyhealth.metrics.binary import binary_metrics_fn
from pyhealth.metrics import multiclass_metrics_fn
import torch.nn.functional as F
from torch.utils.data import DataLoader
import random
from torch.utils.data import Dataset
from tqdm import tqdm
import torch
import os

torch.backends.cuda.matmul.allow_tf32 = True
torch.backends.cudnn.allow_tf32 = True
torch.cuda.set_device(1)


def get_dataset_params(dataset_id):
    return {'type': 'fold', 'range': range(0, 5)}


def get_metric(num_classes, y_true, y_prob):
    """Calculates metrics, handling cases with insufficient data."""
    y_true_np = y_true.cpu().numpy()

    # Edge case: if no samples or only one class in the true labels, metrics are undefined.
    if len(y_true_np) == 0 or len(np.unique(y_true_np)) < 2:
        return 0.0, 0.0, 0.0, 0.0

    if num_classes == 2:
        all_metrics = ["accuracy", "balanced_accuracy", "f1", "roc_auc"]
        y_prob_np = F.softmax(y_prob, dim=-1)[:, 1].cpu().numpy()
        metrics = binary_metrics_fn(y_true_np, y_prob_np, metrics=all_metrics)
        return metrics['accuracy'], metrics['balanced_accuracy'], metrics['f1'], metrics['roc_auc']
    else:
        all_metrics = ["accuracy", "balanced_accuracy", "f1_weighted", "roc_auc_macro_ovo"]
        y_prob_np = F.softmax(y_prob, dim=-1).cpu().numpy()
        metrics = multiclass_metrics_fn(y_true_np, y_prob_np, metrics=all_metrics)
        return metrics['accuracy'], metrics['balanced_accuracy'], metrics['f1_weighted'], metrics['roc_auc_macro_ovo']


def calculate_all_metrics(num_classes, y_true, y_prob, ages, genders, events):
    """
    Calculate metrics for the overall dataset and all specified subgroups.
    """
    results = {}

    # 1. Overall metrics
    results['overall'] = get_metric(num_classes, y_true, y_prob)

    # 2. Define subgroups and their corresponding masks
    subgroups = {
        # Age subgroups
        'age_4.0-4.9': (ages >= 4.0) & (ages < 5.0),
        'age_5.0-5.9': (ages >= 5.0) & (ages < 6.0),
        'age_6.0-6.9': (ages >= 6.0) & (ages < 7.0),
        'age_7.0-7.9': (ages >= 7.0) & (ages < 8.0),
        'age_8.0-8.9': (ages >= 8.0) & (ages < 9.0),
        'age_lt_6.0': ages < 6.0,
        'age_gte_6.0': ages >= 6.0,
        # Gender subgroups
        'gender_female': genders == 0,
        'gender_male': genders == 1,
        # Event subgroups
        'event_0': events == 0,
        'event_1': events == 1,
        'event_2': events == 2,
        'event_3': events == 3,
    }

    # 3. Calculate metrics for each subgroup
    for name, mask in subgroups.items():
        mask = mask.cpu()
        sub_y_true = y_true[mask]
        sub_y_prob = y_prob[mask]
        results[name] = get_metric(num_classes, sub_y_true, sub_y_prob)

    return results


def train(args, model):
    dataloader = DataLoader(dataset=args.train_dataset, batch_size=args.batch_size, shuffle=True, drop_last=True, num_workers=4, pin_memory=True)
    criterion = nn.CrossEntropyLoss()
    model = model.cuda()
    optimizer = torch.optim.AdamW(filter(lambda p: p.requires_grad, model.parameters()), lr=args.lr, weight_decay=0.1)

    # Use a dictionary to store the best metrics for all subgroups
    max_metrics = {}

    no_improve_count = 0
    early_stop_patience = 50

    for epoch in range(args.epochs):
        print(f"Epoch: {epoch}")
        pbar = tqdm(dataloader)
        model.train()
        for _, (features, labels, _, _, _, _) in enumerate(pbar):
            features, labels = features.to(args.device), labels.to(args.device)
            labels_pre = model(features)
            loss = criterion(labels_pre, labels)
            correct = (labels_pre.argmax(1) == labels).type(torch.float).sum().item()
            train_acc = correct / labels.shape[0]
            pbar.set_postfix({'loss': f'{loss:.5f}', 'acc': f'{train_acc:.5f}'})
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()

        # ----------------- Evaluation Step -----------------
        current_metrics_dict = test(args=args, model=model)

        # Initialize max_metrics on the first epoch
        if not max_metrics:
            for subgroup_name, metrics_tuple in current_metrics_dict.items():
                acc, bacc, f1, auc = metrics_tuple
                max_metrics[f"{subgroup_name}_accuracy"] = acc
                max_metrics[f"{subgroup_name}_balanced_accuracy"] = bacc
                max_metrics[f"{subgroup_name}_f1"] = f1
                max_metrics[f"{subgroup_name}_auc"] = auc

        # Check each overall metric independently and update subgroup metrics accordingly
        any_metric_improved = False

        # Unpack current overall metrics for comparison
        current_overall_acc, current_overall_bacc, current_overall_f1, current_overall_auc = current_metrics_dict[
            'overall']

        # 1. Check for ACCURACY improvement
        if max_metrics['overall_accuracy'] <= current_overall_acc:
            any_metric_improved = True
            print(f"Overall Accuracy improved to {current_overall_acc:.4f}. Updating all subgroup accuracies.")
            for subgroup_name, metrics_tuple in current_metrics_dict.items():
                max_metrics[f"{subgroup_name}_accuracy"] = metrics_tuple[0]  # Index 0 is accuracy

        # 2. Check for BALANCED ACCURACY improvement
        if max_metrics['overall_balanced_accuracy'] <= current_overall_bacc:
            any_metric_improved = True
            print(
                f"Overall Balanced Accuracy improved to {current_overall_bacc:.4f}. Updating all subgroup balanced accuracies.")
            for subgroup_name, metrics_tuple in current_metrics_dict.items():
                max_metrics[f"{subgroup_name}_balanced_accuracy"] = metrics_tuple[1]  # Index 1 is balanced_accuracy

        # 3. Check for F1 SCORE improvement
        if max_metrics['overall_f1'] <= current_overall_f1:
            any_metric_improved = True
            print(f"Overall F1 improved to {current_overall_f1:.4f}. Updating all subgroup F1 scores.")
            for subgroup_name, metrics_tuple in current_metrics_dict.items():
                max_metrics[f"{subgroup_name}_f1"] = metrics_tuple[2]  # Index 2 is f1

        # 4. Check for AUC improvement
        if max_metrics['overall_auc'] <= current_overall_auc:
            any_metric_improved = True
            print(f"Overall AUC improved to {current_overall_auc:.4f}. Updating all subgroup AUCs.")
            for subgroup_name, metrics_tuple in current_metrics_dict.items():
                max_metrics[f"{subgroup_name}_auc"] = metrics_tuple[3]  # Index 3 is auc

        # Update early stopping counter
        if any_metric_improved:
            no_improve_count = 0
        else:
            no_improve_count += 1
            print(f"No improvement in any overall metric for {no_improve_count} epochs.")

        print(
            f"Best Overall -> Acc: {max_metrics['overall_accuracy']:.4f}, BACC: {max_metrics['overall_balanced_accuracy']:.4f}, F1: {max_metrics['overall_f1']:.4f}, AUC: {max_metrics['overall_auc']:.4f}")

        if no_improve_count >= early_stop_patience:
            print(f"Early stopping triggered at epoch {epoch} after {early_stop_patience} epochs with no improvement.")
            break

    return max_metrics


def test(args, model):
    model.eval()
    dataloader = DataLoader(args.test_dataset, batch_size=512, num_workers=4, pin_memory=True)

    all_y_true = []
    all_y_prob = []
    all_ages = []
    all_genders = []
    all_events = []

    with torch.no_grad():
        for features, labels, ages, genders, _, events in dataloader:
            features, labels = features.cuda(), labels.cuda()
            preds = model(features)

            all_y_true.append(labels)
            all_y_prob.append(preds)
            all_ages.append(ages)
            all_genders.append(genders)
            all_events.append(events)

    # Concatenate all collected data
    y_true = torch.cat(all_y_true, dim=0)
    y_prob = torch.cat(all_y_prob, dim=0)
    ages = torch.cat(all_ages, dim=0)
    genders = torch.cat(all_genders, dim=0)
    events = torch.cat(all_events, dim=0)

    # Calculate metrics for all subgroups
    all_metrics_results = calculate_all_metrics(args.num_classes, y_true, y_prob, ages, genders, events)

    # Print overall performance for monitoring during training
    overall_accuracy, overall_balanced_accuracy, overall_f1, overall_auc = all_metrics_results['overall']
    print(
        f'Current Test -> Accuracy: {overall_accuracy:.4f}, BACC: {overall_balanced_accuracy:.4f}, F1: {overall_f1:.4f}, AUC: {overall_auc:.4f}')

    return all_metrics_results


def get_dataset(dataset_id=0, seed=42, n=0):
    seed_everything(seed)
    # This function remains unchanged.
    if dataset_id == 0:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/WAV',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=1,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/WAV',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=1,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 1:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/WAV',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=2,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/WAV',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=2,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 2:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/WAV',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=4,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/WAV',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=4,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 3:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/WAV',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=8,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/WAV',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=8,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 4:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/WAV',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=10,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/WAV',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=10,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    if dataset_id == 5:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Children',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=1,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Children',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=1,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 6:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Children',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=2,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Children',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=2,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 7:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Children',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=4,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Children',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=4,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 8:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Children',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=8,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Children',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=8,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 9:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Children',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=10,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Children',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=10,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    if dataset_id == 10:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Adult',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=1,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Adult',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=1,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 11:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Adult',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=2,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Adult',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=2,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 12:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Adult',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=4,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Adult',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=4,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 13:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Adult',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=8,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Adult',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=8,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 14:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Adult',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=10,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Adult',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=10,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 15:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/WAV',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=12,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/WAV',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=12,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 16:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Children',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=12,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Children',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=12,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

    elif dataset_id == 17:
        train_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Adult',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=12,
            fold_index=n,
            mode='train')
        test_dataset = AudioDataset(
            root_dir='/data/chenmingzhi/dataset/Audio/segment/Audio_Adult',
            info_path='/data/chenmingzhi/dataset/Audio/subject_info_with_viq.npz',
            event_classes=[0, 1, 2, 3],
            age_threshold=None,
            age_option=None,
            gender=None,
            duration=12,
            fold_index=n,
            mode='test')
        return train_dataset, test_dataset, 2, 512, 100

def get_model(num_classes=2):
    return AudioClassifier(num_classes=num_classes)



if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    args = parser.parse_args()
    args.device = "cuda"
    dataset_id_list = [15, 16, 17]
    model_id_dict = {
        0: [1], 1: [1], 2: [1], 3: [1], 4: [1], 5: [1], 6: [1],
        7: [1], 8: [1], 9: [1], 10: [1], 11: [1], 12: [1], 13: [1], 14: [1],
        15: [1], 16:[1], 17:[1]
    }

    dataset_id_to_name = {
        0: '1s', 1: '2s', 2: '4s', 3: '8s', 4: '10s', 5: 'Children 1s', 6: 'Children 2s',
        7: 'Children 4s', 8: 'Children 8s', 9: 'Children 10s', 10: 'Adult 1s', 11: 'Adult 2s',
        12: 'Adult 4s', 13: 'Adult 8s', 14: 'Adult 10s', 15: '12s', 16: 'Children 12s', 17: 'Adult 12s'
    }
    dataset_id_to_num_classes = {id: 2 for id in dataset_id_list}
    model_id_to_name = {0: 'AudioTransformer'}

    # Define the order of subgroups and metrics for consistent Excel reporting
    subgroup_names = [
        'overall', 'age_4.0-4.9', 'age_5.0-5.9', 'age_6.0-6.9', 'age_7.0-7.9', 'age_8.0-8.9',
        'age_lt_6.0', 'age_gte_6.0', 'gender_female', 'gender_male',
        'event_0', 'event_1', 'event_2', 'event_3'
    ]
    metric_names = ['accuracy', 'balanced_accuracy', 'f1', 'auc']

    filename = '/data/Result/Result.xlsx'
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
        workbook = Workbook()
        if 'Sheet' in workbook.sheetnames: del workbook['Sheet']

    for dataset_id in dataset_id_list:
        dataset_name = dataset_id_to_name[dataset_id]
        num_classes = dataset_id_to_num_classes[dataset_id]
        params_info = get_dataset_params(dataset_id)
        param_range = params_info['range']

        if dataset_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(title=dataset_name)
            # Dynamically create headers
            headers = ['model_name', 'fold_seed']
            for s_name in subgroup_names:
                for m_name in metric_names:
                    headers.append(f"{s_name}_{m_name}")
            worksheet.append(headers)
        else:
            worksheet = workbook[dataset_name]

        for model_id in model_id_dict.get(dataset_id, []):
            model_name = model_id_to_name[model_id]
            for n in param_range:  # n is the fold index
                existing = any(
                    row[0].value == model_name and row[1].value == n for row in worksheet.iter_rows(min_row=2))
                if existing:
                    print(f"Skipping existing record: {dataset_name}, {model_name}, fold {n}")
                    continue

                print(f"\n--- Running: Dataset={dataset_name}, Model={model_name}, Fold={n} ---")

                try:
                    train_dataset, test_dataset, num_classes_ds, batch_size, epochs = get_dataset(dataset_id=dataset_id,
                                                                                                  n=n)
                    assert num_classes == num_classes_ds, "Class number mismatch!"
                except Exception as e:
                    print(f"Failed to get dataset: {e}")
                    continue

                args.num_classes = num_classes
                args.batch_size = batch_size
                args.epochs = epochs
                args.train_dataset = train_dataset
                args.test_dataset = test_dataset
                args.lr = 2e-4

                try:
                    model = get_model(num_classes)
                    model.to(args.device)
                except Exception as e:
                    print(f"Failed to initialize model: {e}")
                    continue

                try:
                    final_metrics = train(args, model)
                except Exception as e:
                    print(f"Training failed: {e}")
                    continue

                # Construct row data in the correct order
                row_data = [model_name, n]
                for s_name in subgroup_names:
                    for m_name in metric_names:
                        key = f"{s_name}_{m_name}"
                        row_data.append(final_metrics.get(key, 0.0))  # Use .get for safety

                worksheet.append(row_data)
                workbook.save(filename)
                print(f"Saved results for: {dataset_name}, {model_name}, fold {n}")

    print("\nAll experiments completed!")